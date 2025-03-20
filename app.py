import os
import logging
from datetime import datetime
from flask import Flask, request, jsonify, send_file
from openpyxl import load_workbook
from flask_cors import CORS

# Configuración de la aplicación Flask
app = Flask(__name__)
CORS(app)  # Habilitar CORS para permitir solicitudes desde tu frontend

# Configuración del logging
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# Ruta del archivo original de Excel
PLANTILLA_EXCEL = 'PARTES DE TRABAJO PARA PYTHON.xlsx'

# Constantes para el cálculo de horas
HORAS_STANDARD = 7.75  # 7 horas y 45 minutos
HORAS_LIMITE_EXTRAS = 10.0  # Límite para considerar horas extras

# Contraseña de acceso
CONTRASENA_CORRECTA = "2579"

# Almacenamiento de sesiones de usuario (en una aplicación real, usarías una base de datos)
user_sessions = {}

@app.route('/')
def index():
    return "API de registro de partes de trabajo funcionando"

@app.route('/api/start', methods=['POST'])
def start_conversation():
    """Inicia una nueva conversación de registro de partes"""
    session_id = request.json.get('session_id', datetime.now().strftime("%Y%m%d%H%M%S"))
    
    # Inicializar una nueva sesión para el usuario
    if session_id not in user_sessions:
        user_sessions[session_id] = {
            'estado': 'VERIFICAR_CONTRASENA',
            'respuestas': {},
            'proyectos': [],
            'horas_acumuladas': 0.0,
            'proyecto_actual': None
        }
    
    return jsonify({
        'session_id': session_id,
        'mensaje': 'Por favor, introduce la contraseña para continuar:',
        'opciones': None,
        'estado': 'VERIFICAR_CONTRASENA'
    })

@app.route('/api/message', methods=['POST'])
def process_message():
    """Procesa cada mensaje del usuario según el estado actual de la conversación"""
    data = request.json
    session_id = data.get('session_id')
    mensaje = data.get('mensaje')
    
    if not session_id or session_id not in user_sessions:
        return jsonify({'error': 'Sesión no válida'}), 400
    
    session = user_sessions[session_id]
    estado = session['estado']
    
    # Procesar el mensaje según el estado actual
    if estado == 'VERIFICAR_CONTRASENA':
        return verificar_contrasena(session, mensaje)
    elif estado == 'NOMBRE':
        return procesar_nombre(session, mensaje)
    elif estado == 'TIPO_TRABAJO':
        return procesar_tipo_trabajo(session, mensaje)
    elif estado == 'ORDEN_TRABAJO':
        return procesar_orden_trabajo(session, mensaje)
    elif estado == 'NUM_PARTE':
        return procesar_num_parte(session, mensaje)
    elif estado == 'PARTE_CERRADO':
        return procesar_parte_cerrado(session, mensaje)
    elif estado == 'HORAS_PROYECTO':
        return procesar_horas_proyecto(session, mensaje)
    elif estado == 'OTRO_PROYECTO':
        return procesar_otro_proyecto(session, mensaje)
    elif estado == 'HORAS_TOTALES':
        return procesar_horas_totales(session, mensaje)
    else:
        return jsonify({'error': 'Estado no reconocido'}), 400

def verificar_contrasena(session, contrasena):
    """Verifica si la contraseña introducida es correcta."""
    if contrasena == CONTRASENA_CORRECTA:
        session['estado'] = 'NOMBRE'
        return jsonify({
            'mensaje': '¡Contraseña correcta! Voy a ayudarte a completar el parte diario. ¿Cuál es tu nombre?',
            'opciones': None,
            'estado': 'NOMBRE'
        })
    else:
        return jsonify({
            'mensaje': 'Contraseña errónea. Por favor, intenta de nuevo:',
            'opciones': None,
            'estado': 'VERIFICAR_CONTRASENA'
        })

def procesar_nombre(session, nombre):
    """Procesa el nombre del usuario y solicita el tipo de trabajo."""
    session['respuestas']['NOMBRE'] = nombre
    
    # Obtener automáticamente la fecha actual
    fecha_actual = datetime.now()
    session['respuestas']['Nº DIA'] = str(fecha_actual.day)
    
    # Nombres de los meses en español
    meses = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", 
             "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    session['respuestas']['MES'] = meses[fecha_actual.month - 1]
    
    session['estado'] = 'TIPO_TRABAJO'
    return jsonify({
        'mensaje': f'Gracias, {nombre}.\n¿El trabajo es facturable, una orden de trabajo o no facturable?',
        'opciones': ['Facturable', 'Orden de Trabajo', 'No Facturable'],
        'estado': 'TIPO_TRABAJO'
    })

def procesar_tipo_trabajo(session, tipo_trabajo):
    """Procesa el tipo de trabajo y determina el siguiente paso según la respuesta."""
    # Crear un nuevo diccionario para el proyecto actual
    proyecto_actual = {}
    
    # Si el trabajo es una orden de trabajo, preguntamos qué orden de trabajo es
    if tipo_trabajo == 'Orden de Trabajo':
        session['es_orden_trabajo'] = True
        session['estado'] = 'ORDEN_TRABAJO'
        
        return jsonify({
            'mensaje': '¿Qué orden de trabajo es? (Por favor, especifica el nombre o número de la orden)',
            'opciones': None,
            'estado': 'ORDEN_TRABAJO'
        })
    else:
        # Si es Facturable o No Facturable, continuamos con el flujo normal
        proyecto_actual['FACTURABLE O ORDEN DE TRABAJO'] = tipo_trabajo
        session['proyecto_actual'] = proyecto_actual
        
        # Si el trabajo es No Facturable, saltamos las preguntas de parte
        if tipo_trabajo == 'No Facturable':
            # Establecemos valores predeterminados para las preguntas que saltamos
            proyecto_actual['Nº DE PARTE'] = 'No aplica'
            proyecto_actual['PARTE CERRADO'] = 'No aplica'
            
            session['estado'] = 'HORAS_PROYECTO'
            return jsonify({
                'mensaje': '¿Cuántas horas has dedicado a este proyecto? (Formato: 2.5 para 2 horas y 30 minutos)',
                'opciones': None,
                'estado': 'HORAS_PROYECTO'
            })
        else:
            # Si es Facturable, preguntamos por el número de parte
            session['estado'] = 'NUM_PARTE'
            return jsonify({
                'mensaje': 'Introduce el número de parte. Si no aplica, escribe "No aplica".',
                'opciones': None,
                'estado': 'NUM_PARTE'
            })

def procesar_orden_trabajo(session, orden_trabajo):
    """Procesa la orden de trabajo específica."""
    # Crear un nuevo diccionario para el proyecto actual si no existe
    if 'proyecto_actual' not in session or not session['proyecto_actual']:
        session['proyecto_actual'] = {}
    
    # Guardamos el tipo de trabajo como la orden de trabajo específica
    session['proyecto_actual']['FACTURABLE O ORDEN DE TRABAJO'] = orden_trabajo
    
    session['estado'] = 'NUM_PARTE'
    return jsonify({
        'mensaje': 'Introduce el número de parte. Si no aplica, escribe "No aplica".',
        'opciones': None,
        'estado': 'NUM_PARTE'
    })

def procesar_num_parte(session, num_parte):
    """Procesa el número de parte y determina el siguiente paso según la respuesta."""
    session['proyecto_actual']['Nº DE PARTE'] = num_parte
    
    # Si el número de parte es "No aplica", saltamos la pregunta de parte cerrado
    if num_parte.lower() == 'no aplica':
        session['proyecto_actual']['PARTE CERRADO'] = 'No aplica'
        
        session['estado'] = 'HORAS_PROYECTO'
        return jsonify({
            'mensaje': '¿Cuántas horas has dedicado a este proyecto? (Formato: 2.5 para 2 horas y 30 minutos)',
            'opciones': None,
            'estado': 'HORAS_PROYECTO'
        })
    else:
        # Si hay un número de parte válido, preguntamos si está cerrado
        session['estado'] = 'PARTE_CERRADO'
        return jsonify({
            'mensaje': '¿El parte está cerrado?',
            'opciones': ['Sí', 'No'],
            'estado': 'PARTE_CERRADO'
        })

def procesar_parte_cerrado(session, parte_cerrado):
    """Procesa si el parte está cerrado y solicita las horas dedicadas al proyecto."""
    session['proyecto_actual']['PARTE CERRADO'] = parte_cerrado
    
    session['estado'] = 'HORAS_PROYECTO'
    return jsonify({
        'mensaje': '¿Cuántas horas has dedicado a este proyecto? (Formato: 2.5 para 2 horas y 30 minutos)',
        'opciones': None,
        'estado': 'HORAS_PROYECTO'
    })

def procesar_horas_proyecto(session, texto_horas):
    """Procesa las horas del proyecto y pregunta si hay más proyectos."""
    try:
        # Validar que sea un número decimal
        horas = float(texto_horas.replace(',', '.'))
        session['proyecto_actual']['TOTAL DE HORAS'] = str(horas)
        
        # Actualizar el total de horas acumuladas
        if 'horas_acumuladas' not in session:
            session['horas_acumuladas'] = 0.0
        session['horas_acumuladas'] += horas
        
        # Añadir el proyecto actual a la lista de proyectos
        session['proyectos'].append(session['proyecto_actual'])
        
        session['estado'] = 'OTRO_PROYECTO'
        return jsonify({
            'mensaje': '¿Has trabajado en algún otro proyecto hoy?',
            'opciones': ['Sí', 'No'],
            'estado': 'OTRO_PROYECTO'
        })
    except ValueError:
        return jsonify({
            'mensaje': 'Por favor, introduce un número válido (ejemplo: 2.5 para 2 horas y 30 minutos).',
            'opciones': None,
            'estado': 'HORAS_PROYECTO'
        })

def procesar_otro_proyecto(session, respuesta):
    """Procesa si hay más proyectos y determina el siguiente paso."""
    if respuesta.lower() == 'sí' or respuesta.lower() == 'si':
        # Si hay más proyectos, volvemos a preguntar por el tipo de trabajo
        # Limpiar el proyecto actual
        session['proyecto_actual'] = None
        session.pop('es_orden_trabajo', None)
        
        session['estado'] = 'TIPO_TRABAJO'
        return jsonify({
            'mensaje': '¿El siguiente trabajo es facturable, una orden de trabajo o no facturable?',
            'opciones': ['Facturable', 'Orden de Trabajo', 'No Facturable'],
            'estado': 'TIPO_TRABAJO'
        })
    else:
        # Si no hay más proyectos, preguntamos por las horas totales
        horas_sugeridas = session.get('horas_acumuladas', 0.0)
        
        session['estado'] = 'HORAS_TOTALES'
        return jsonify({
            'mensaje': f'¿Cuántas horas totales has trabajado hoy? (Formato: 7.75 para 7 horas y 45 minutos)\nHoras sugeridas: {horas_sugeridas:.2f} (suma de todas las horas de proyectos)',
            'opciones': None,
            'estado': 'HORAS_TOTALES'
        })

def procesar_horas_totales(session, texto_horas):
    """Procesa las horas totales, calcula automáticamente las horas de bolsa y extras, y completa el archivo Excel."""
    try:
        # Validar que sea un número decimal
        horas_totales = float(texto_horas.replace(',', '.'))
        session['respuestas']['HORAS TOTALES'] = str(horas_totales)
        
        # Calcular automáticamente las horas de bolsa y extras
        horas_bolsa = 0
        horas_extras = 0
        
        if horas_totales > HORAS_STANDARD:
            if horas_totales <= HORAS_LIMITE_EXTRAS:
                # Horas de bolsa (entre 7h45m y 10h)
                horas_bolsa = round(horas_totales - HORAS_STANDARD, 2)
            else:
                # Horas de bolsa (hasta 10h) + horas extras (más de 10h)
                horas_bolsa = round(HORAS_LIMITE_EXTRAS - HORAS_STANDARD, 2)
                horas_extras = round(horas_totales - HORAS_LIMITE_EXTRAS, 2)
                
        session['respuestas']['HORAS BOLSA'] = str(horas_bolsa)
        session['respuestas']['HORAS EXTRAS'] = str(horas_extras)
        
        try:
            # Generar el archivo Excel con las respuestas y los proyectos
            nombre_archivo = generar_excel(session['respuestas'], session['proyectos'])
            
            # Crear un resumen de los datos
            resumen = "✅ Resumen de los datos introducidos:\n\n"
            campos_importantes = ['NOMBRE', 'Nº DIA', 'MES', 
                                'HORAS TOTALES', 'HORAS BOLSA', 'HORAS EXTRAS']
            
            for campo in campos_importantes:
                if campo in session['respuestas']:
                    resumen += f"• {campo}: {session['respuestas'][campo]}\n"
            
            resumen += "\nProyectos registrados:\n"
            for i, proyecto in enumerate(session['proyectos'], 1):
                resumen += f"\nProyecto {i}:\n"
                resumen += f"• Tipo: {proyecto['FACTURABLE O ORDEN DE TRABAJO']}\n"
                resumen += f"• Nº de parte: {proyecto['Nº DE PARTE']}\n"
                resumen += f"• Parte cerrado: {proyecto['PARTE CERRADO']}\n"
                resumen += f"• Horas: {proyecto['TOTAL DE HORAS']}\n"
            
            # Almacenar el nombre del archivo en la sesión para su descarga
            session['archivo_excel'] = nombre_archivo
            session['resumen'] = resumen
            session['estado'] = 'COMPLETADO'
            
            return jsonify({
                'mensaje': f'Procesando tu solicitud...\n\nHoras totales trabajadas: {horas_totales} horas\nHoras regulares: {min(horas_totales, HORAS_STANDARD)} horas\nHoras de bolsa: {horas_bolsa} horas\nHoras extras: {horas_extras} horas\n\n{resumen}\n\n¡Gracias! El archivo Excel ha sido completado.',
                'opciones': None,
                'estado': 'COMPLETADO',
                'archivo_disponible': True,
                'resumen': resumen
            })
            
        except Exception as e:
            logger.error(f"Error al procesar el archivo Excel: {e}")
            return jsonify({
                'mensaje': 'Ha ocurrido un error al procesar el archivo Excel. Por favor, intenta de nuevo.',
                'opciones': None,
                'estado': 'ERROR'
            })
    except ValueError:
        return jsonify({
            'mensaje': 'Por favor, introduce un número válido.',
            'opciones': None,
            'estado': 'HORAS_TOTALES'
        })

@app.route('/api/download/<session_id>', methods=['GET'])
def download_excel(session_id):
    """Permite descargar el archivo Excel generado"""
    if session_id not in user_sessions or 'archivo_excel' not in user_sessions[session_id]:
        return jsonify({'error': 'Archivo no disponible'}), 404
    
    archivo = user_sessions[session_id]['archivo_excel']
    
    # Enviar el archivo para descarga
    return send_file(archivo, as_attachment=True)

@app.route('/api/reset/<session_id>', methods=['POST'])
def reset_session(session_id):
    """Reinicia la sesión para un nuevo parte de trabajo"""
    if session_id in user_sessions:
        # Eliminar el archivo Excel si existe
        if 'archivo_excel' in user_sessions[session_id]:
            try:
                os.remove(user_sessions[session_id]['archivo_excel'])
            except:
                pass
        
        # Limpiar la sesión
        user_sessions.pop(session_id, None)
    
    return jsonify({'mensaje': 'Sesión reiniciada correctamente'})

# Función para generar una copia del archivo Excel con las respuestas
def generar_excel(respuestas, proyectos):
    """
    Genera una copia del archivo Excel original y la completa con las respuestas y proyectos.
    Retorna el nombre del archivo generado.
    """
    # Generar un nombre único para el archivo que incluya el nombre del trabajador y la fecha
    fecha_actual = datetime.now().strftime("%d%m%Y")
    nombre_usuario = respuestas['NOMBRE'].replace(' ', '_')
    nombre_archivo = f"PARTE_TRABAJO_{nombre_usuario}_{fecha_actual}.xlsx"
    
    # Cargar el archivo de Excel original
    wb = load_workbook(PLANTILLA_EXCEL)
    sheet = wb.active
    
    # Reemplazar los valores en las celdas para información general
    campos_generales = ['NOMBRE', 'Nº DIA', 'MES', 'HORAS TOTALES', 'HORAS BOLSA', 'HORAS EXTRAS']
    
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None and isinstance(cell.value, str):
                # Reemplazar campos generales
                for campo in campos_generales:
                    marcador = '{' + campo + '}'
                    if marcador in cell.value and campo in respuestas:
                        cell.value = cell.value.replace(marcador, respuestas.get(campo, ''))
    
    # Buscar la fila de la plantilla que contiene los marcadores para los proyectos
    fila_plantilla = None
    
    for row_idx, row in enumerate(sheet.iter_rows(), 1):
        for cell in row:
            if cell.value is not None and isinstance(cell.value, str) and '{FACTURABLE O ORDEN DE TRABAJO}' in cell.value:
                fila_plantilla = row_idx
                break
        if fila_plantilla:
            break
    
    if fila_plantilla:
        # Determinar la estructura de la fila de la plantilla
        estructura_fila = []
        for col_idx, cell in enumerate(sheet[fila_plantilla], 1):
            if cell.value is not None and isinstance(cell.value, str):
                for campo in ['FACTURABLE O ORDEN DE TRABAJO', 'Nº DE PARTE', 'PARTE CERRADO', 'TOTAL DE HORAS']:
                    marcador = '{' + campo + '}'
                    if marcador in cell.value:
                        estructura_fila.append((col_idx, campo))
                        break
        
        # Si tenemos más de un proyecto, necesitamos duplicar la fila de la plantilla para cada proyecto adicional
        if len(proyectos) > 1:
            # Duplicar la fila de la plantilla para cada proyecto adicional
            for i in range(1, len(proyectos)):
                # Insertar una nueva fila después de la fila de la plantilla
                sheet.insert_rows(fila_plantilla + i)
                
                # Copiar el formato y estructura de la fila de la plantilla
                for col_idx in range(1, sheet.max_column + 1):
                    # Copiar el valor, formato y estilo de la celda de la plantilla
                    source_cell = sheet.cell(row=fila_plantilla, column=col_idx)
                    target_cell = sheet.cell(row=fila_plantilla + i, column=col_idx)
                    
                    # Copiar el valor
                    target_cell.value = source_cell.value
        
        # Ahora rellenar los datos de cada proyecto en su fila correspondiente
        for i, proyecto in enumerate(proyectos):
            fila_actual = fila_plantilla + i
            
            # Rellenar los campos del proyecto en la fila actual
            for col_idx, campo in estructura_fila:
                celda = sheet.cell(row=fila_actual, column=col_idx)
                if celda.value is not None and isinstance(celda.value, str):
                    marcador = '{' + campo + '}'
                    if marcador in celda.value:
                        celda.value = celda.value.replace(marcador, proyecto.get(campo, ''))
                    else:
                        celda.value = proyecto.get(campo, '')
    
    # Guardar el archivo Excel modificado con un nuevo nombre
    wb.save(nombre_archivo)
    
    return nombre_archivo

# Ejecutar la aplicación
if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
